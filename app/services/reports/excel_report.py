"""
Excel report generator - 3 sheets with rich formatting.
Sheet 1: Summary Dashboard
Sheet 2: Student Results (sorted by roll number)
Sheet 3: Ranked List (sorted by marks)
"""

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.schemas.extraction import StudentRecord, AnalyticsSummary

# Colors
INDIGO = "6366F1"
INDIGO_LIGHT = "E0E7FF"
EMERALD = "10B981"
EMERALD_LIGHT = "D1FAE5"
ROSE = "F43F5E"
ROSE_LIGHT = "FFE4E6"
AMBER = "F59E0B"
AMBER_LIGHT = "FEF3C7"
GOLD = "F59E0B"
SILVER = "9CA3AF"
BRONZE = "B45309"
WHITE = "FFFFFF"
GRAY_50 = "F9FAFB"
GRAY_100 = "F3F4F6"
GRAY_700 = "374151"
GRAY_900 = "111827"

THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=INDIGO, end_color=INDIGO, fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, color=GRAY_900, size=14)
SUBTITLE_FONT = Font(name="Calibri", bold=True, color=INDIGO, size=12)
BODY_FONT = Font(name="Calibri", size=10, color=GRAY_700)
PASS_FILL = PatternFill(start_color=EMERALD_LIGHT, end_color=EMERALD_LIGHT, fill_type="solid")
FAIL_FILL = PatternFill(start_color=ROSE_LIGHT, end_color=ROSE_LIGHT, fill_type="solid")
GOLD_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
SILVER_FILL = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
BRONZE_FILL = PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid")


def generate_excel_report(
    students: list[StudentRecord],
    analytics: AnalyticsSummary,
    university_name: str = "",
    subject_name: str = "",
) -> bytes:
    wb = Workbook()

    _create_summary_sheet(wb, analytics, university_name, subject_name)
    _create_results_sheet(wb, students)
    _create_ranked_sheet(wb, students)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _create_summary_sheet(
    wb: Workbook,
    analytics: AnalyticsSummary,
    university_name: str,
    subject_name: str,
):
    ws = wb.active
    ws.title = "Summary Dashboard"
    ws.sheet_properties.tabColor = INDIGO

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "ResultKraft - Exam Result Analysis"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Subtitle
    ws.merge_cells("A2:H2")
    ws["A2"] = f"{university_name} | {subject_name}" if subject_name else university_name
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    # Stats grid
    row = 4
    stats = [
        ("Total Students", analytics.total_students, INDIGO_LIGHT),
        ("Pass Count", analytics.pass_count, EMERALD_LIGHT),
        ("Fail Count", analytics.fail_count, ROSE_LIGHT),
        ("Pass %", f"{analytics.pass_percentage}%", EMERALD_LIGHT),
        ("Class Average", analytics.class_average, INDIGO_LIGHT),
        ("Highest Score", analytics.highest_score, "FEF3C7"),
        ("Lowest Score", analytics.lowest_score, ROSE_LIGHT),
        ("Median", analytics.median_score, INDIGO_LIGHT),
        ("Std Deviation", analytics.std_deviation, GRAY_100),
    ]

    col = 1
    for label, value, color in stats:
        ws.cell(row=row, column=col, value=label).font = Font(bold=True, size=9, color=GRAY_700)
        ws.cell(row=row, column=col).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=col).border = THIN_BORDER

        val_cell = ws.cell(row=row + 1, column=col, value=value)
        val_cell.font = Font(bold=True, size=14, color=GRAY_900)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        val_cell.border = THIN_BORDER

        ws.column_dimensions[get_column_letter(col)].width = 16
        col += 1
        if col > 5:
            col = 1
            row += 3

    # Grade distribution
    grade_row = row + 3
    ws.cell(row=grade_row, column=1, value="Grade Distribution").font = SUBTITLE_FONT
    grade_row += 1

    grade_headers = ["Grade", "Count", "Percentage"]
    for i, h in enumerate(grade_headers):
        cell = ws.cell(row=grade_row, column=i + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    total = analytics.total_students or 1
    for grade, count in sorted(analytics.grade_distribution.items()):
        grade_row += 1
        ws.cell(row=grade_row, column=1, value=grade).border = THIN_BORDER
        ws.cell(row=grade_row, column=2, value=count).border = THIN_BORDER
        pct_cell = ws.cell(row=grade_row, column=3, value=round(count / total * 100, 1))
        pct_cell.number_format = '0.0"%"'
        pct_cell.border = THIN_BORDER

    # Score ranges
    range_row = grade_row + 2
    ws.cell(row=range_row, column=1, value="Score Ranges").font = SUBTITLE_FONT
    range_row += 1

    for h_i, h in enumerate(["Range", "Count"]):
        cell = ws.cell(row=range_row, column=h_i + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for rng, count in analytics.score_ranges.items():
        range_row += 1
        ws.cell(row=range_row, column=1, value=rng).border = THIN_BORDER
        ws.cell(row=range_row, column=2, value=count).border = THIN_BORDER


def _create_results_sheet(wb: Workbook, students: list[StudentRecord]):
    ws = wb.create_sheet("Student Results")
    ws.sheet_properties.tabColor = EMERALD

    headers = [
        "Roll No", "Name", "Father's Name", "Enrollment No",
        "Subject", "External", "Internal", "Total", "Grade", "Status"
    ]
    widths = [14, 24, 24, 18, 28, 12, 12, 10, 8, 10]

    for i, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    # Sort by roll number for this sheet
    sorted_students = sorted(students, key=lambda s: s.roll_no)

    for row_idx, student in enumerate(sorted_students, 2):
        data = [
            student.roll_no, student.student_name, student.father_name or "",
            student.enrollment_no or "", student.subject_name or "",
            student.ext_marks, student.int_marks, student.total_marks,
            student.grade, student.pass_fail,
        ]

        # Determine row color
        if student.rank_in_class and student.rank_in_class <= 3:
            fill = GOLD_FILL
        elif student.pass_fail == "PASS":
            fill = PASS_FILL
        else:
            fill = FAIL_FILL

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx > 4 else "left")

    # Auto-filter and freeze panes
    ws.auto_filter.ref = f"A1:J{len(sorted_students) + 1}"
    ws.freeze_panes = "A2"


def _create_ranked_sheet(wb: Workbook, students: list[StudentRecord]):
    ws = wb.create_sheet("Ranked List")
    ws.sheet_properties.tabColor = GOLD

    headers = [
        "Rank", "Name", "Roll No", "Total Marks",
        "External", "Internal", "Grade", "Percentile", "Status"
    ]
    widths = [8, 24, 14, 12, 12, 12, 8, 12, 10]

    for i, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(i)].width = width

    # Already sorted by total marks descending
    ranked = sorted(students, key=lambda s: s.total_marks, reverse=True)

    for row_idx, student in enumerate(ranked, 2):
        rank = row_idx - 1

        # Medal colors for top 3
        if rank == 1:
            fill = GOLD_FILL
        elif rank == 2:
            fill = SILVER_FILL
        elif rank == 3:
            fill = BRONZE_FILL
        elif student.pass_fail == "PASS":
            fill = PASS_FILL
        else:
            fill = FAIL_FILL

        data = [
            rank, student.student_name, student.roll_no,
            student.total_marks, student.ext_marks, student.int_marks,
            student.grade, student.percentile or 0, student.pass_fail,
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
